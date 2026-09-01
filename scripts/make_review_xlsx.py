# -*- coding: utf-8 -*-
"""make_review_xlsx.py — 生成10轮100场盲测复盘xlsx"""
import csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = list(csv.DictReader(open('match_data/summary_100.csv', encoding='utf-8-sig')))

wb = Workbook()
ws = wb.active
ws.title = '10轮100场复盘'
headers = ['轮', '对阵', '赛果', '实际方向', '主占%', '主冷热', '客占%', '客冷热', '盲测判定', '判定规则', '命中']
ws.append(headers)
header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(color='FFFFFF', bold=True, size=11)
for c in range(1, len(headers)+1):
    cell = ws.cell(1, c)
    cell.fill = header_fill; cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
green = PatternFill('solid', fgColor='C6EFCE')
red = PatternFill('solid', fgColor='FFC7CE')
for r in rows:
    ws.append([r['轮'], r['对阵'], r['赛果'], r['方向'], r['主占%'], r['主冷热'], r['客占%'], r['客冷热'], r['盲测方向'], r['判定规则'], r['命中']])
    row_idx = ws.max_row
    if r['命中'] == '1':
        ws.cell(row_idx, 11).fill = green
    else:
        ws.cell(row_idx, 11).fill = red
# 统计行
tot_hit = sum(1 for r in rows if r['命中'] == '1')
ws.append(['总计', '', '', '', '', '', '', '', f'{tot_hit}/100', '', ''])
# 轮次汇总
ws2 = wb.create_sheet('轮次汇总')
ws2.append(['轮次', '主胜', '平', '客胜', '盲测命中', '主胜率'])
from collections import defaultdict
for rnd in range(1, 11):
    rr = [r for r in rows if int(r['轮']) == rnd]
    h = sum(1 for r in rr if r['方向'] == '主胜')
    d = sum(1 for r in rr if r['方向'] == '平')
    a = sum(1 for r in rr if r['方向'] == '客胜')
    hit = sum(1 for r in rr if r['命中'] == '1')
    ws2.append([f'第{rnd}轮', h, d, a, f'{hit}/10', f'{h*10:.0f}%'])
# 机构操盘特点表
ws3 = wb.create_sheet('机构操盘特点')
ws3.append(['维度', '发现', '结论'])
ws3.append(['欧赔方向预测力', '前10轮100场各机构"下调方向vs赛果"命中率普遍20-30%', '欧赔单一方向信号弱，不能独立使用'])
ws3.append(['欧赔最活跃(下调主胜多)', '利*23场、18B*23场、易**14场、Crow*13场', '利*/18B*倾向表达"看好主队"立场'])
ws3.append(['欧赔最保守', '立*仅12场下调主胜、澳*16场、马*18场', '立*/澳*欧赔调整少，信号含金量需看幅度'])
ws3.append(['亚盘方向预测力', '100场各机构"盘口升降vs赛果"命中率30-38%', '亚盘方向优于欧赔，但需结合水位'])
ws3.append(['亚盘最活跃(升盘最多)', '威廉**35场升盘、利*32场、伟*30场、明*28场', '威廉/利*/伟*升盘信号较活跃'])
ws3.append(['亚盘最保守', '立*89场持平、马*74场、In*****66场、澳*60场', '立*几乎不调整盘口(6%命中=无预测价值),马*/In*****保守'])
ws3.append(['亚盘方向命中最高', '利*38%、1X***38%、威廉**37%、明*34%、伟*34%', '利*/威廉/明*的升盘=较强看好主队信号'])
ws3.append(['亚盘方向命中最低', '立*6%、In*****11%、马*11%、澳*19%', '这几家亚盘调整参考价值低'])
ws3.append(['整体方向信号强度', '欧赔20-30% vs 亚盘30-38%', '亚盘方向信号强于欧赔约1.3-1.5倍'])
wb.save('match_data/复盘_10轮100场.xlsx')
print('已生成 match_data/复盘_10轮100场.xlsx')
print('命中率', tot_hit, '/100 =', tot_hit/len(rows)*100, '%')
