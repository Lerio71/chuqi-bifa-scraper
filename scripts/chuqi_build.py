# -*- coding: utf-8 -*-
"""
chuqi_build.py — 从 structured.json 数据驱动生成比赛数据工作簿（7个sheet）
自动提取：赛程赛况 / 阵容数据 / 对阵分析 / 数据统计 / 指数(欧赔·亚盘·必发)
无需人工整理，纯脚本生成。
用法：python chuqi_build.py <match_id> [out.xlsx]
"""
import json, os, sys, re
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="4472C4")
SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SEC_FONT = Font(bold=True, size=12, color="1F4E79")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

def bt(ms):
    """毫秒时间戳 → 北京时间字符串"""
    if not ms: return ""
    try:
        return (datetime.fromtimestamp(ms / 1000, tz=timezone.utc) + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""

def write_table(ws, start_row, headers, rows, col_widths=None):
    r = start_row
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
    r += 1
    for row in rows:
        for j, v in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = WRAP; c.border = BORDER
        r += 1
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return r

def section(ws, row, text, span=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SEC_FONT; c.fill = SEC_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1

def fv(obj):
    """{fields, values} → [{field: value}]"""
    if not isinstance(obj, dict) or "fields" not in obj: return []
    fields, values = obj["fields"], obj.get("values", [])
    return [dict(zip(fields, v)) for v in values]

# ---- 事件类型映射 ----
EVT = {1: "进球", 3: "黄牌", 4: "红牌", 9: "换人", 2: "攻防事件"}
POSN = {1: "主队", 2: "客队"}

# ---- 技术统计 kind 映射 ----
STATKIND = {1: "进球", 11: "黄牌", 20: "角球", 50: "控球率%", 61: "进攻", 62: "危险进攻", 121: "射正", 122: "射偏"}

def clean_html(s):
    return re.sub(r"<[^>]+>", "", str(s)).strip()

def supplier_map(sup):
    if isinstance(sup, dict):
        return {str(k): v.get("name", str(k)) for k, v in sup.items()}
    if isinstance(sup, list):
        return {str(s["id"]): s.get("name", "") for s in sup}
    return {}

def is_std_key(k):
    """仅标准胜平负/亚盘盘（纯数字key），跳过 '51-1'/'-1' 等让球/大小球子盘"""
    return re.fullmatch(r"\d+", str(k)) is not None

def main(mid, outpath=None):
    base = os.path.join(os.getcwd(), "match_data", mid)
    struct = json.load(open(os.path.join(base, "structured.json"), encoding="utf-8"))
    B = struct["boards"]
    info = B["detail"]["info"]
    home, away = info["home"], info["away"]
    title = f"{info.get('match','')} 第{info.get('round','')}轮 | {home} vs {away} | {bt(info.get('scheduletime'))} | 全场 {info.get('home_score')}:{info.get('away_score')}"
    wb = Workbook()

    # ============ Sheet1 赛程赛况 ============
    ws = wb.active; ws.title = "赛程赛况"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT; ws.merge_cells("A1:F1")
    r = section(ws, 3, "一、基本信息")
    rows = [["赛事", f"{info.get('match','')} 第{info.get('round','')}轮", "比赛时间", bt(info.get("scheduletime"))],
            ["主队", f"{home}（主）", "客队", f"{away}（客）"],
            ["全场比分", f"{info.get('home_score')}:{info.get('away_score')}", "半场比分", f"{info.get('home_half_score')}:{info.get('away_half_score')}"],
            ["比赛结果", "主胜" if info.get("home_score",0) > info.get("away_score",0) else ("客胜" if info.get("home_score",0) < info.get("away_score",0) else "平局"), "球场", info.get("stadium","") or "—"],
            ["天气", info.get("weather",""), "气温", info.get("temperature","")]]
    r = write_table(ws, r, ["项目", "内容", "项目", "内容"], rows, [14, 40, 14, 40]) + 1

    # 技术统计
    r = section(ws, r, "二、技术统计（主队-客队）")
    stat_rows = []
    for v in B["lineup"]["data"]["stat"]["values"]:
        kind = v[1]
        if kind in STATKIND and v[0] == 0:
            stat_rows.append([STATKIND[kind], v[2], v[3]])
    rows = [[home, a, k, b, away] for k, a, b in stat_rows]
    r = write_table(ws, r, ["主队", "主队值", "指标", "客队值", "客队"], rows, [16, 12, 14, 12, 16]) + 1

    # 事件：进球 / 黄牌 / 换人
    inc = B["detail"].get("eventTrend", {}).get("incidents", [])
    goals, cards, subs, others = [], [], [], []
    for e in inc:
        t = e.get("time", ""); ty = e.get("type"); pos = POSN.get(e.get("position"), "")
        label = EVT.get(ty, f"事件{ty}")
        if ty == 1: goals.append([t, "进球", pos, "", f"{pos}进球"])
        elif ty == 3: cards.append([t, pos, "黄牌"])
        elif ty == 9: subs.append([t, pos, "换人"])
        else: others.append([t, pos, label])
    r = section(ws, r, "三、进球")
    r = write_table(ws, r, ["时间", "事件", "球队", "方式", "比分"], goals, [10, 10, 14, 22, 14]) + 1
    r = section(ws, r, "四、换人")
    r = write_table(ws, r, ["时间", "球队", "事件"], subs, [10, 12, 20]) + 1
    r = section(ws, r, "五、黄牌")
    r = write_table(ws, r, ["时间", "球队", "事件"], cards, [10, 12, 20]) + 1
    r = section(ws, r, "六、事件时间轴（含攻防事件）")
    rows = [[t, pos, label] for t, pos, label in others]
    r = write_table(ws, r, ["时间", "球队", "事件"], rows, [10, 12, 20]) + 1
    ws.freeze_panes = "A3"

    # ============ Sheet2 阵容数据 ============
    ws = wb.create_sheet("阵容数据")
    ld = B["lineup"]["data"]
    rem = ld.get("remarks", {})
    fmt = ld.get("formatio", {}).get("values", [[None, None]])
    ws.cell(row=1, column=1, value=f"阵容数据 | {title}").font = TITLE_FONT; ws.merge_cells("A1:F1")
    r = section(ws, 3, "一、球队总览")
    rows = [["全队身价(€)", f"{rem.get('home_totalworth','')}万", f"{rem.get('away_totalworth','')}万", "全队身价(€)"],
            ["平均年龄", f"{rem.get('home_averageage','')}岁", f"{rem.get('away_averageage','')}岁", "平均年龄"],
            ["主教练", rem.get("home_coach", ""), rem.get("away_coach", ""), "主教练"],
            ["阵型", fmt[0][0] if fmt and fmt[0] else "", fmt[0][1] if fmt and fmt[0] else "", "阵型"],
            ["首发身价(€)", f"{rem.get('home_firstworth','')}万", f"{rem.get('away_firstworth','')}万", "首发身价(€)"]]
    rows = [[home, a, b, away] for a, b, *_ in rows]
    rows = [[home, o[1], o[2], away] for o in [
        ["", rem.get("home_totalworth",""), rem.get("away_totalworth",""), ""],
        ["", f"{rem.get('home_averageage','')}岁", f"{rem.get('away_averageage','')}岁", ""],
        ["", rem.get("home_coach",""), rem.get("away_coach",""), ""],
        ["", fmt[0][0] if fmt and fmt[0] else "", fmt[0][1] if fmt and fmt[0] else "", ""],
        ["", f"{rem.get('home_firstworth','')}万", f"{rem.get('away_firstworth','')}万", ""]]]
    r = write_table(ws, r, ["主队", "主队值", "客队值", "客队"], rows, [14, 24, 24, 14]) + 1

    # 首发/替补（group: 1=主 2=客；flags: 1=首发）
    lineup = ld["lineup"]
    lv = fv(lineup)
    posmap = {90: "守门员", 80: "后卫", 70: "中场", 60: "前锋", 50: "替补"}
    for g, gname in [(1, home), (2, away)]:
        members = [x for x in lv if x.get("group") == g]
        start = [x for x in members if x.get("flags") == 1]
        bench = [x for x in members if x.get("flags") != 1]
        f = fmt[0][0] if fmt and fmt[0] else ""
        f2 = fmt[0][1] if fmt and fmt[0] else ""
        r = section(ws, r, f"二、首发阵容（{gname} {f if g==1 else f2}）")
        rows = [[x.get("number",""), x.get("name",""), posmap.get(x.get("position"), x.get("position",""))] for x in start]
        r = write_table(ws, r, ["号码", "球员", "位置"], rows, [10, 28, 12]) + 1
        r = section(ws, r, f"三、替补阵容（{gname}）")
        rows = [[x.get("number",""), x.get("name",""), posmap.get(x.get("position"), x.get("position",""))] for x in bench]
        r = write_table(ws, r, ["号码", "球员", "位置"], rows, [10, 28, 12]) + 1

    # 伤停
    r = section(ws, r, "四、伤停阵容")
    inj_rows = []
    for x in ld.get("injury", {}).get("list", []):
        g = "主队" if x.get("group") == 1 else "客队"
        inj_rows.append([x.get("title", x.get("name","")), g, x.get("reason",""), x.get("status","")])
    r = write_table(ws, r, ["球员", "球队", "原因", "状态"], inj_rows, [24, 12, 20, 12]) + 1
    ws.freeze_panes = "A3"

    # ============ Sheet3 对阵分析 ============
    ws = wb.create_sheet("对阵分析")
    fd = B["fenxi"]["data"]
    team = {t[0]: t[1] for t in fd["team"]["values"]} if isinstance(fd["team"].get("values"), list) else {}
    ws.cell(row=1, column=1, value=f"对阵分析 | {title}").font = TITLE_FONT; ws.merge_cells("A1:F1")
    homeid, awayid = info.get("homeid"), info.get("awayid")

    recs = fv(fd["records"])
    def rowfmt(x):
        h = team.get(x.get("homeid"), x.get("homeid")); a = team.get(x.get("awayid"), x.get("awayid"))
        return [bt(x.get("scheduletime")), x.get("competition",""), h,
                f"{x.get('home_score')}-{x.get('away_score')}({x.get('home_score_half')}-{x.get('away_score_half')})", a,
                x.get("home_corner",""), x.get("away_corner","")]
    # 历史交锋：双方互相对战
    head2head = [x for x in recs if x.get("homeid") in (homeid, awayid) and x.get("awayid") in (homeid, awayid)]
    r = section(ws, 3, "一、历史交锋（双方互相对战）")
    rows = [rowfmt(x) for x in head2head[:10]]
    r = write_table(ws, r, ["日期", "赛事", "主", "比分(半场)", "客", "主队角球", "客队角球"], rows, [12, 10, 14, 18, 14, 10, 10]) + 1

    # 近期战绩
    def recent(teamid):
        return [x for x in recs if x.get("homeid") == teamid or x.get("awayid") == teamid][:10]
    for tid, tname in [(homeid, home), (awayid, away)]:
        r = section(ws, r, f"二、近期战绩（{tname} 近10场）")
        rows = [rowfmt(x) for x in recent(tid)]
        r = write_table(ws, r, ["日期", "赛事", "主", "比分(半场)", "客", "主队角球", "客队角球"], rows, [12, 10, 14, 18, 14, 10, 10]) + 1

    # 未来比赛
    r = section(ws, r, "三、未来比赛")
    fut = []
    for tid, tname in [(homeid, home), (awayid, away)]:
        for x in fv(fd.get(f"{'home' if tid==homeid else 'away'}_schedule", {})):
            opp = team.get(x.get("homeid") if tid != x.get("homeid") else x.get("awayid"), "?")
            fut.append([tname, x.get("competition",""), bt(x.get("scheduletime")), "主" if x.get("homeid")==tid else "客", opp])
    r = write_table(ws, r, ["球队", "赛事", "日程", "主/客", "对手"], fut, [10, 10, 16, 10, 16]) + 1
    ws.freeze_panes = "A3"

    # ============ Sheet4 数据统计（积分榜） ============
    ws = wb.create_sheet("数据统计")
    ws.cell(row=1, column=1, value=f"数据统计（联赛积分榜） | {title}").font = TITLE_FONT; ws.merge_cells("A1:H1")
    ints = B["stats"].get("integrals")
    if isinstance(ints, dict) and "0" in ints: ints = ints["0"]
    rows = []
    if isinstance(ints, list):
        # 每队可能多维度（主/客/总），去重保留场次最全的一条
        best = {}
        for x in ints:
            tid = x.get("teamid")
            matches = (x.get("wincount") or 0) + (x.get("drawcount") or 0) + (x.get("losecount") or 0)
            cur = best.get(tid)
            if cur is None or matches > cur[0]:
                best[tid] = (matches, x)
        for _, x in sorted(best.values(), key=lambda kv: kv[1].get("rank") or 999):
            rows.append([x.get("rank",""), x.get("alias", x.get("teamid","")), x.get("wincount",""), x.get("drawcount",""),
                         x.get("losecount",""), x.get("getscore",""), x.get("losescore",""), x.get("integral","")])
    r = write_table(ws, 3, ["排名", "球队", "胜", "平", "负", "进球", "失球", "积分"], rows, [8, 22, 8, 8, 8, 8, 8, 8])
    ws.freeze_panes = "A4"

    # ============ Sheet5 指数-欧赔 ============
    ws = wb.create_sheet("指数-欧赔")
    ws.cell(row=1, column=1, value=f"欧赔（胜平负） | {title}").font = TITLE_FONT; ws.merge_cells("A1:J1")
    o = B["odds1x2"]["data"]
    sup = supplier_map(o["supplier"])
    r = section(ws, 3, "一、各公司初赔→即时赔")
    rows = []
    for sid, od in o["odds"].items():
        if not is_std_key(sid):
            continue
        cp, js = od.get("cp", {}), od.get("js", {})
        rows.append([sup.get(str(sid), sid), cp.get("a",""), cp.get("b",""), cp.get("c",""), bt(cp.get("time")),
                     js.get("a",""), js.get("b",""), js.get("c",""), bt(js.get("time"))])
    r = write_table(ws, r, ["公司", "初主胜", "初客胜", "初平局", "初盘时间", "即主胜", "即客胜", "即平局", "即时时间"], rows, [12, 10, 10, 10, 16, 10, 10, 10, 16]) + 1

    r = section(ws, r, "二、20家欧赔变化（开盘→赛前）")
    # 收集所有公司变化
    for name in struct["changes"]:
        ch = struct["changes"][name]
        spf = ch.get("spf")
        if not spf or not spf["rows"]: continue
        rows = []
        for row in spf["rows"]:
            status = clean_html(row[-1])
            if "滚" in status: continue  # 排除赛内
            rows.append([name, row[1], row[2], row[3], row[4], row[5], row[7], status])
        r = write_table(ws, r, ["公司", "赛况", "比分", "胜", "平", "负", "变化时间", "状态"], rows, [12, 10, 8, 10, 10, 10, 16, 8]) + 1
    ws.freeze_panes = "A3"

    # ============ Sheet6 指数-亚盘 ============
    ws = wb.create_sheet("指数-亚盘")
    ws.cell(row=1, column=1, value=f"亚盘（让球盘） | {title}").font = TITLE_FONT; ws.merge_cells("A1:J1")
    o = B["oddsah"]["data"]
    sup = supplier_map(o["supplier"])
    r = section(ws, 3, "一、各公司初盘→即时盘")
    rows = []
    for sid, od in o["odds"].items():
        if not is_std_key(sid):
            continue
        cp, js = od.get("cp", {}), od.get("js", {})
        rows.append([sup.get(str(sid), sid), cp.get("a",""), cp.get("b",""), cp.get("c",""), bt(cp.get("time")),
                     js.get("a",""), js.get("b",""), js.get("c",""), bt(js.get("time"))])
    r = write_table(ws, r, ["公司", "初主水", "初客水", "初让球", "初盘时间", "即主水", "即客水", "即让球", "即时时间"], rows, [12, 10, 10, 10, 16, 10, 10, 10, 16]) + 1

    r = section(ws, r, "二、20家亚盘变化（开盘→赛前）")
    for name in struct["changes"]:
        ch = struct["changes"][name]
        rq = ch.get("rq")
        if not rq or not rq["rows"]: continue
        rows = []
        for row in rq["rows"]:
            status = clean_html(row[-1])
            if "滚" in status: continue
            rows.append([name, row[1], row[2], row[3], row[4], row[5], row[7], status])
        r = write_table(ws, r, ["公司", "赛况", "比分", "主水", "盘口", "客水", "变化时间", "状态"], rows, [12, 10, 8, 10, 10, 10, 16, 8]) + 1
    ws.freeze_panes = "A3"

    # ============ Sheet7 指数-必发 ============
    ws = wb.create_sheet("指数-必发")
    ws.cell(row=1, column=1, value=f"必发交易指数 | {title}").font = TITLE_FONT; ws.merge_cells("A1:H1")
    bf = B["bifa"].get("allData", [])
    r = section(ws, 3, "一、必发成交汇总（胜平负）")
    rows = []
    for x in bf:
        s = x.get("summary", {})
        rows.append([x.get("name",""), s.get("odds",""), s.get("amount",""), s.get("per",""),
                     s.get("profit",""), s.get("payout",""), s.get("hot","")])
    r = write_table(ws, r, ["选项", "必发赔率", "成交量", "占比", "盈亏指数", "赔付率", "冷热指数"], rows, [10, 12, 16, 10, 12, 10, 10]) + 1

    r = section(ws, r, "二、必发成交时间序列（最近10个时间点）")
    rows = []
    for x in bf:
        for p in x.get("echart", [])[:10]:
            rows.append([x.get("name",""), bt(p.get("time")), p.get("amount","")])
    r = write_table(ws, r, ["选项", "时间", "成交额"], rows, [10, 16, 16]) + 1
    ws.freeze_panes = "A3"

    if not outpath:
        outpath = os.path.join(base, f"{mid}_比赛数据.xlsx")
    wb.save(outpath)
    print("已生成:", outpath)
    return outpath

if __name__ == "__main__":
    mid = sys.argv[1] if len(sys.argv) > 1 else "14586460"
    out = sys.argv[2] if len(sys.argv) > 2 else None
    main(mid, out)
